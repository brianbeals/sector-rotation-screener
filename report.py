"""Excel and HTML report writers."""
from __future__ import annotations

import math
import os
from datetime import date, datetime
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import config

NAVY    = "1E3A5F"
ACCENT  = "2E86C1"
CARD    = "D6EAF8"
BG      = "F4F6F9"
WIN     = "1E8449"
RISK    = "C0392B"
CAUTION = "D4AC0D"


# --- Excel --------------------------------------------------------------------

def write_excel(rows: List[Dict],
                cycle_info: Dict,
                vintage_info: Dict,
                out_path: str,
                backtest_df: Optional[pd.DataFrame] = None,
                backtest_summary: Optional[Dict] = None) -> str:
    wb = Workbook()

    # Sheet 1: Sector Screen
    ws = wb.active
    ws.title = "Sector Screen"

    # Header info row above the grid
    ws.cell(row=1, column=1, value=f"Sector Rotation Screen — {date.today():%B %d, %Y}").font = \
        Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    vintage_str = (
        f"Generated {vintage_info.get('generated_at','')} · "
        f"Prices through {vintage_info.get('prices_through','')} · "
        f"FRED vintage {vintage_info.get('fred_vintage','')}"
    )
    ws.cell(row=2, column=1, value=vintage_str).font = Font(italic=True, size=10, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)

    headers = [
        "Sector ETF", "Name", "Last", "% From 52W High",
        "Seasonality (n yrs)", "Cycle Fit", "Rel Strength",
        "RS 1m vs SPY", "RS 3m vs SPY", "RS 6m vs SPY",
        "Composite", "Signal",
    ]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 32

    thin = Side(border_style="thin", color="CCCCCC")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, row in enumerate(rows, header_row + 1):
        season_label = f"{row['seasonality_score']:.1f} ({row.get('seasonality_n', 0)})"
        if row.get("seasonality_thin"):
            season_label += " ⚠"
        ws.cell(row=r, column=1,  value=row["ticker"])
        ws.cell(row=r, column=2,  value=row["name"])
        ws.cell(row=r, column=3,  value=row["last_price"]).number_format = "$#,##0.00"
        ws.cell(row=r, column=4,  value=row["pct_from_52w_high"]).number_format = "0.00%"
        ws.cell(row=r, column=5,  value=season_label)
        ws.cell(row=r, column=6,  value=row["cycle_fit_score"]).number_format = "0.0"
        ws.cell(row=r, column=7,  value=row["rs_score"]).number_format = "0.0"
        ws.cell(row=r, column=8,  value=row["rs_1m"]).number_format = "0.00%"
        ws.cell(row=r, column=9,  value=row["rs_3m"]).number_format = "0.00%"
        ws.cell(row=r, column=10, value=row["rs_6m"]).number_format = "0.00%"
        ws.cell(row=r, column=11, value=row["composite"]).number_format = "0.0"
        sig_cell = ws.cell(row=r, column=12, value=row["signal"])
        color = {"Buy": WIN, "Avoid": RISK}.get(row["signal"], CAUTION)
        sig_cell.font = Font(bold=True, color="FFFFFF")
        sig_cell.fill = PatternFill("solid", fgColor=color)
        sig_cell.alignment = Alignment(horizontal="center")
        if (r - header_row) % 2 == 0:
            for c in range(1, 12):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=CARD)
        for c in range(1, 13):
            ws.cell(row=r, column=c).border = box

    widths = [11, 24, 9, 13, 18, 11, 13, 13, 13, 13, 11, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{header_row + 1}"

    # Sheet 2: Cycle Context
    ws2 = wb.create_sheet("Cycle Context")
    ws2["A1"] = "Cycle Phase Classification"
    ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws2.merge_cells("A1:B1")
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 28

    inp = cycle_info["inputs"]
    ctx_rows = [
        ("Phase (active)",         cycle_info["phase"]),
        ("Algorithm phase",        cycle_info.get("algo_phase", cycle_info["phase"])),
        ("Override active",        "Yes" if cycle_info.get("override_active") else "No"),
        ("Reasoning",              cycle_info.get("why", "")),
        ("10Y yield",              inp.get("DGS10")),
        ("2Y yield",               inp.get("DGS2")),
        ("10Y - 2Y spread",        inp.get("spread")),
        ("Spread 6m change",       inp.get("spread_6m_chg")),
        ("INDPRO YoY %",           inp.get("INDPRO_YoY")),
        ("INDPRO 3m change",       inp.get("INDPRO_3m_chg")),
        ("10Y as of",              str(inp.get("DGS10_date") or "")),
        ("INDPRO as of",           str(inp.get("INDPRO_date") or "")),
        ("Composite weights",      f"Seasonality {config.WEIGHTS.seasonality:.0%} / "
                                    f"Cycle {config.WEIGHTS.cycle_fit:.0%} / "
                                    f"RS {config.WEIGHTS.rel_strength:.0%}"),
        ("MAX_POSITIONS",          config.MAX_POSITIONS),
        ("MIN_SCORE_TO_HOLD",      config.MIN_SCORE_TO_HOLD),
        ("TRADE_COST_BPS",         config.TRADE_COST_BPS),
        ("TAXABLE_ACCOUNT",        config.TAXABLE_ACCOUNT),
    ]
    for i, (k, v) in enumerate(ctx_rows, 3):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True, color="1A1A2A")
        cell = ws2.cell(row=i, column=2, value=v)
        if isinstance(v, float):
            cell.number_format = "0.00"
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 70

    # Sheet 3: Backtest
    if backtest_df is not None and not backtest_df.empty:
        wsb = wb.create_sheet("Backtest")
        # Summary block
        wsb["A1"] = f"Backtest Summary — last {config.BACKTEST_YEARS} years"
        wsb["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        wsb["A1"].fill = PatternFill("solid", fgColor=NAVY)
        wsb.merge_cells("A1:F1")
        wsb["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        wsb.row_dimensions[1].height = 26

        s = backtest_summary or {}
        summary_rows = [
            ("Strategy cumulative return", _pct(s.get("strategy_cum"))),
            ("SPY cumulative return",      _pct(s.get("spy_cum"))),
            ("Strategy CAGR",              _pct(s.get("strategy_cagr"))),
            ("SPY CAGR",                   _pct(s.get("spy_cagr"))),
            ("Strategy Sharpe (ann)",      _num(s.get("strategy_sharpe"))),
            ("SPY Sharpe (ann)",           _num(s.get("spy_sharpe"))),
            ("Strategy max drawdown",      _pct(s.get("strategy_maxdd"))),
            ("SPY max drawdown",           _pct(s.get("spy_maxdd"))),
            ("Months",                     int(s.get("months", 0))),
            ("Beats SPY net of cost",      "Yes" if s.get("beats_spy_net") else "NO"),
        ]
        for i, (k, v) in enumerate(summary_rows, 3):
            wsb.cell(row=i, column=1, value=k).font = Font(bold=True, color="1A1A2A")
            wsb.cell(row=i, column=2, value=v)
        if not s.get("beats_spy_net"):
            warn = wsb.cell(row=3 + len(summary_rows), column=1,
                            value="Strategy did NOT beat SPY net of cost over this window.")
            warn.font = Font(bold=True, color="FFFFFF")
            warn.fill = PatternFill("solid", fgColor=RISK)
            wsb.merge_cells(start_row=3 + len(summary_rows), start_column=1,
                            end_row=3 + len(summary_rows), end_column=6)

        # Monthly table starts a few rows below
        start_row = 3 + len(summary_rows) + 3
        cols = ["Date", "Phase", "Holdings", "Strategy Ret", "SPY Ret",
                "Turnover", "Cost Drag", "Strategy Equity", "SPY Equity"]
        for c, h in enumerate(cols, 1):
            cell = wsb.cell(row=start_row, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for j, (idx, row) in enumerate(backtest_df.iterrows(), start_row + 1):
            wsb.cell(row=j, column=1, value=str(idx))
            wsb.cell(row=j, column=2, value=row["phase"])
            wsb.cell(row=j, column=3, value=row["holdings"])
            wsb.cell(row=j, column=4, value=row["strategy_ret"]).number_format = "0.00%"
            wsb.cell(row=j, column=5, value=row["spy_ret"]).number_format = "0.00%"
            wsb.cell(row=j, column=6, value=row["turnover"]).number_format = "0.00"
            wsb.cell(row=j, column=7, value=row["cost_drag"]).number_format = "0.00%"
            wsb.cell(row=j, column=8, value=row["strategy_equity"]).number_format = "0.000"
            wsb.cell(row=j, column=9, value=row["spy_equity"]).number_format = "0.000"
            if j % 2 == 0:
                for c in range(1, 10):
                    wsb.cell(row=j, column=c).fill = PatternFill("solid", fgColor=CARD)

        widths_b = [12, 14, 28, 13, 13, 11, 12, 16, 14]
        for i, w in enumerate(widths_b, 1):
            wsb.column_dimensions[get_column_letter(i)].width = w
        wsb.freeze_panes = f"A{start_row + 1}"

    # File props
    wb.properties.creator = "Brian Beals"
    wb.properties.lastModifiedBy = "Beals, Brian"
    wb.properties.category = "Investing Research"
    wb.properties.title = f"Sector Rotation Screen — {date.today().isoformat()}"

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


def _pct(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    return f"{v*100:+.2f}%"

def _num(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    return f"{v:.2f}"


# --- HTML ---------------------------------------------------------------------

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Sector Rotation Screen — {{ run_date }}</title>
<style>
  :root {
    --navy:#1E3A5F; --accent:#2E86C1; --card:#D6EAF8;
    --bg:#F4F6F9; --body:#1A1A2A;
    --win:#1E8449; --risk:#C0392B; --caution:#D4AC0D;
  }
  * { box-sizing: border-box; }
  body { margin: 0; padding: 0; background: var(--bg); color: var(--body);
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    font-size: 15px; }
  .wrap { max-width: 1180px; margin: 0 auto; padding: 24px; }
  header.bar {
    background: var(--navy); color: #fff; padding: 18px 24px; border-radius: 8px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
  }
  header.bar h1 { margin: 0 0 4px 0; font-size: 22px; font-weight: 600; }
  header.bar .meta { font-size: 13px; opacity: 0.85; }
  .banner {
    margin-top: 12px; padding: 12px 16px; border-radius: 6px;
    font-weight: 600; font-size: 14px;
  }
  .banner.warn { background: var(--risk); color: #fff; }
  .banner.tax  { background: var(--caution); color: #1A1A2A; }
  .banner.note { background: #fff; color: #1A1A2A;
                 border-left: 4px solid var(--accent); }
  .phase-card {
    background: #fff; border-radius: 8px; padding: 18px 22px; margin-top: 16px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    display: grid; grid-template-columns: auto 1fr; gap: 18px; align-items: center;
  }
  .phase-badge {
    display: inline-block; padding: 8px 16px; border-radius: 999px;
    background: var(--accent); color: #fff; font-weight: 600; font-size: 16px;
    min-width: 130px; text-align: center;
  }
  .phase-badge.override { background: var(--caution); color: #1A1A2A; }
  .phase-why { color: #333; line-height: 1.5; }
  .phase-inputs { margin-top: 14px; font-size: 13px; color: #444;
    display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; }
  .phase-inputs .label { color: #777; font-size: 11px; text-transform: uppercase;
    letter-spacing: 0.5px; }
  .phase-inputs .val { color: var(--navy); font-size: 17px; font-weight: 600; }
  .phase-inputs .delta { color: #555; font-size: 12px; }
  h2.section {
    background: var(--navy); color: #fff; padding: 10px 16px; margin: 22px 0 12px 0;
    border-radius: 6px; font-size: 16px; font-weight: 600;
  }
  table { width: 100%; border-collapse: collapse; background: #fff;
          border-radius: 8px; overflow: hidden;
          box-shadow: 0 2px 8px rgba(0,0,0,0.06); }
  th { background: var(--accent); color: #fff; padding: 10px 8px; text-align: left;
       cursor: pointer; user-select: none; font-size: 13px; font-weight: 600; }
  th.num, td.num { text-align: right; }
  th:hover { background: #2575a8; }
  td { padding: 9px 8px; border-bottom: 1px solid #eef2f6; font-size: 14px; }
  tr:nth-child(even) td { background: #fbfcfd; }
  tr:hover td { background: var(--card); }
  .signal { display: inline-block; padding: 3px 10px; border-radius: 4px;
            font-weight: 600; font-size: 12px; color: #fff; }
  .sig-Buy { background: var(--win); }
  .sig-Hold { background: var(--caution); color: #1A1A2A; }
  .sig-Avoid { background: var(--risk); }
  .pill { display: inline-block; padding: 3px 10px; border-radius: 4px;
          font-weight: 600; font-size: 12px; color: #fff; }
  .pill-buy { background: var(--win); }
  .pill-hold { background: var(--caution); color: #1A1A2A; }
  .pill-avoid { background: var(--risk); }
  .thin-flag { color: var(--caution); font-weight: 600; cursor: help; }
  .heatmap { width: 100%; border-collapse: collapse; background: #fff;
             box-shadow: 0 2px 8px rgba(0,0,0,0.06); border-radius: 8px; overflow: hidden; }
  .heatmap th, .heatmap td { padding: 8px 6px; text-align: center;
                              font-size: 12px; border-bottom: 1px solid #eef2f6; }
  .heatmap th { background: var(--accent); color: #fff; }
  .heatmap td.row-label { text-align: left; font-weight: 600; color: var(--navy);
                          background: #f7faff; }
  .legend { font-size: 12px; color: #555; margin-top: 6px; }
  .rs-bars { background: #fff; border-radius: 8px; padding: 16px;
             box-shadow: 0 2px 8px rgba(0,0,0,0.06); }
  .rs-row { display: grid; grid-template-columns: 80px 1fr 80px;
            align-items: center; gap: 10px; padding: 4px 0; font-size: 13px; }
  .rs-bar { background: #eef2f6; border-radius: 3px; height: 18px; position: relative; }
  .rs-bar .fill { position: absolute; top: 0; bottom: 0; }
  .rs-bar .fill.pos { background: var(--win); left: 50%; }
  .rs-bar .fill.neg { background: var(--risk); right: 50%; }
  .rs-bar .axis { position: absolute; left: 50%; top: 0; bottom: 0; width: 1px;
                  background: rgba(0,0,0,0.2); }
  .equity-card {
    background: #fff; border-radius: 8px; padding: 16px 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
  }
  .stats-grid { display: grid; grid-template-columns: repeat(4, 1fr);
                gap: 10px; margin-bottom: 12px; }
  .stat { background: #f7faff; border-radius: 6px; padding: 10px 12px; }
  .stat .lbl { color: #777; font-size: 11px; text-transform: uppercase;
               letter-spacing: 0.5px; }
  .stat .val { color: var(--navy); font-size: 17px; font-weight: 700;
               margin-top: 4px; }
  .stat .sub { color: #555; font-size: 12px; margin-top: 2px; }
  footer { text-align: center; color: #888; font-size: 12px; margin: 28px 0 12px;
           line-height: 1.6; }
</style>
</head>
<body>
<div class="wrap">

<header class="bar">
  <h1>Sector Rotation Screen</h1>
  <div class="meta">
    {{ run_date }} · 11 SPDR sector ETFs vs SPY · Composite weights:
    Seasonality {{ w_season }} / Cycle {{ w_cycle }} / RS {{ w_rs }}
  </div>
</header>

{% if backtest_warning %}
<div class="banner warn">
  ⚠ Strategy did NOT beat SPY net of {{ trade_cost_bps }} bps trading cost over the {{ backtest_years }}-year backtest. Cumulative: strategy {{ strategy_cum }} vs SPY {{ spy_cum }}.
</div>
{% endif %}

{% if taxable_account %}
<div class="banner tax">
  Taxable-account mode: monthly rotation generates short-term gains. Strategy must beat SPY by your marginal-rate-vs-LTCG differential to be worth it net of tax.
</div>
{% endif %}

<div class="phase-card">
  <div class="phase-badge {{ 'override' if override_active else '' }}">{{ phase }}</div>
  <div>
    <div class="phase-why">{{ phase_why }}</div>
    <div class="phase-inputs">
      <div>
        <div class="label">10Y - 2Y spread</div>
        <div class="val">{{ spread }}</div>
        <div class="delta">6m change: {{ spread_6m_chg }}</div>
      </div>
      <div>
        <div class="label">10Y yield</div>
        <div class="val">{{ dgs10 }}%</div>
        <div class="delta">2Y: {{ dgs2 }}%</div>
      </div>
      <div>
        <div class="label">INDPRO YoY</div>
        <div class="val">{{ indpro_yoy }}%</div>
        <div class="delta">3m trend: {{ indpro_3m_chg }}</div>
      </div>
      <div>
        <div class="label">INDPRO as of</div>
        <div class="val" style="font-size: 14px;">{{ indpro_date }}</div>
        <div class="delta">10Y as of: {{ dgs10_date }}</div>
      </div>
    </div>
  </div>
</div>

<h2 class="section">Sector Scores &amp; Signals</h2>
<table id="scores">
  <thead>
    <tr>
      <th data-sort="text">ETF</th>
      <th data-sort="text">Name</th>
      <th class="num" data-sort="num">Last</th>
      <th class="num" data-sort="num">% From 52W High</th>
      <th class="num" data-sort="num">Seasonality</th>
      <th class="num" data-sort="num">n yrs</th>
      <th class="num" data-sort="num">Cycle Fit</th>
      <th class="num" data-sort="num">Rel Strength</th>
      <th class="num" data-sort="num">RS 3m</th>
      <th class="num" data-sort="num">Composite</th>
      <th data-sort="text">Signal</th>
    </tr>
  </thead>
  <tbody>
  {% for r in rows %}
    <tr>
      <td>{{ r.ticker }}</td>
      <td>{{ r.name }}</td>
      <td class="num">${{ "%.2f"|format(r.last_price) }}</td>
      <td class="num">{{ "%.2f%%"|format(100*r.pct_from_52w_high) }}</td>
      <td class="num">{{ "%.1f"|format(r.seasonality_score) }}{% if r.seasonality_thin %} <span class="thin-flag" title="Thin sample: fewer than {{ trust_yrs }} years of history">⚠</span>{% endif %}</td>
      <td class="num">{{ r.seasonality_n }}</td>
      <td class="num">{{ "%.1f"|format(r.cycle_fit_score) }}</td>
      <td class="num">{{ "%.1f"|format(r.rs_score) }}</td>
      <td class="num">{{ "%.2f%%"|format(100*r.rs_3m) }}</td>
      <td class="num"><b>{{ "%.1f"|format(r.composite) }}</b></td>
      <td><span class="signal sig-{{ r.signal }}">{{ r.signal }}</span></td>
    </tr>
  {% endfor %}
  </tbody>
</table>
<div class="legend">Click any column header to sort. ⚠ flags sectors with under {{ trust_yrs }} years of history. Composite weights and signal thresholds tunable in <code>config.py</code>.</div>

{% if has_backtest %}
<h2 class="section">Backtest — last {{ backtest_years }} years</h2>
<div class="equity-card">
  <div class="stats-grid">
    <div class="stat"><div class="lbl">Strategy cum return</div><div class="val">{{ strategy_cum }}</div><div class="sub">SPY: {{ spy_cum }}</div></div>
    <div class="stat"><div class="lbl">Strategy CAGR</div><div class="val">{{ strategy_cagr }}</div><div class="sub">SPY: {{ spy_cagr }}</div></div>
    <div class="stat"><div class="lbl">Strategy Sharpe</div><div class="val">{{ strategy_sharpe }}</div><div class="sub">SPY: {{ spy_sharpe }}</div></div>
    <div class="stat"><div class="lbl">Max drawdown</div><div class="val">{{ strategy_maxdd }}</div><div class="sub">SPY: {{ spy_maxdd }}</div></div>
  </div>
  {{ equity_svg|safe }}
  <div class="legend">Equity curves start at 1.00. Trading cost {{ trade_cost_bps }} bps applied to turnover at each rebalance. Cash held when no sectors clear MIN_SCORE_TO_HOLD={{ min_score }}.</div>
</div>
{% else %}
<div class="banner note">No backtest run this session — call <code>screener.run_screen(skip_backtest=False)</code> to include it.</div>
{% endif %}

<h2 class="section">Seasonality Heatmap (avg monthly return by sector)</h2>
<table class="heatmap">
  <thead>
    <tr>
      <th>Sector</th>
      {% for m in months %}<th>{{ m }}</th>{% endfor %}
    </tr>
  </thead>
  <tbody>
  {% for tk, cells in heatmap %}
    <tr>
      <td class="row-label">{{ tk }}</td>
      {% for cell in cells %}
      <td style="background:{{ cell.color }}; color:{{ cell.text_color }};">{{ cell.label }}</td>
      {% endfor %}
    </tr>
  {% endfor %}
  </tbody>
</table>
<div class="legend">Green &gt; 0%, red &lt; 0%. Cells require ≥{{ min_years }} years of history for that month.</div>

<h2 class="section">3-Month Relative Strength vs SPY</h2>
<div class="rs-bars">
  {% for r in rs_rows %}
    <div class="rs-row">
      <div><b>{{ r.ticker }}</b></div>
      <div class="rs-bar">
        <div class="axis"></div>
        {% if r.rs_3m >= 0 %}
        <div class="fill pos" style="width: {{ r.bar_pct }}%;"></div>
        {% else %}
        <div class="fill neg" style="width: {{ r.bar_pct }}%;"></div>
        {% endif %}
      </div>
      <div class="num" style="text-align:right;">{{ "%+.2f%%"|format(100*r.rs_3m) }}</div>
    </div>
  {% endfor %}
</div>

<footer>
  Prepared by <a href="https://brianbeals.com" target="_blank" style="color:inherit;">Brian Beals</a> · {{ run_date_long }}<br>
  Price data through {{ prices_through }} · FRED vintage {{ fred_vintage }}<br>
  © {{ current_year }} Brian Beals · <a href="https://github.com/brianbeals/sector-rotation-screener" target="_blank" style="color:inherit;">github.com/brianbeals/sector-rotation-screener</a>
</footer>

</div>

<script>
document.querySelectorAll("#scores th").forEach((th, i) => {
  th.addEventListener("click", () => {
    const tbody = th.closest("table").querySelector("tbody");
    const rows = Array.from(tbody.querySelectorAll("tr"));
    const isNum = th.dataset.sort === "num";
    const dir = th.dataset.dir === "asc" ? "desc" : "asc";
    document.querySelectorAll("#scores th").forEach(t => t.dataset.dir = "");
    th.dataset.dir = dir;
    rows.sort((a, b) => {
      let av = a.children[i].innerText.trim();
      let bv = b.children[i].innerText.trim();
      if (isNum) {
        av = parseFloat(av.replace(/[^0-9.\-]/g, "")) || 0;
        bv = parseFloat(bv.replace(/[^0-9.\-]/g, "")) || 0;
        return dir === "asc" ? av - bv : bv - av;
      }
      return dir === "asc" ? av.localeCompare(bv) : bv.localeCompare(av);
    });
    rows.forEach(r => tbody.appendChild(r));
  });
});
</script>

{% if drilldown %}
{% for parent_tk, dd in drilldown.items() %}
<section style="margin-top:2rem;">
<h2>Drill-Down: {{ parent_tk }} ({{ sector_names.get(parent_tk, parent_tk) }})
{% if dd.parent_signal == 'Buy' %}
  <span class="pill pill-buy" style="font-size:14px; vertical-align:middle; margin-left:8px;">BUY</span>
{% else %}
  <span class="pill pill-hold" style="font-size:14px; vertical-align:middle; margin-left:8px;">WATCH</span>
{% endif %}
</h2>

{% if dd.subsectors %}
<h3 style="margin-top:1rem;">Sub-Sector ETFs <span class="dim">(ranked by composite)</span></h3>
<table class="data" style="margin-top:.5rem;">
<thead><tr>
<th>ETF</th><th>Theme</th><th data-sort="num">Composite</th><th>Signal</th>
<th data-sort="num">RSI</th><th data-sort="num">vs 20d</th><th data-sort="num">vs 50d</th>
<th data-sort="num">Vol</th><th>Entry Timing</th>
</tr></thead>
<tbody>
{% for s in dd.subsectors %}
<tr>
<td><strong>{{ s.ticker }}</strong></td>
<td>{{ s.name }}</td>
<td style="font-weight:600; color:{% if s.signal == 'Buy' %}#27ae60{% elif s.signal == 'Avoid' %}#c0392b{% else %}inherit{% endif %};">{{ "%.1f"|format(s.composite) }}</td>
<td><span class="pill {% if s.signal == 'Buy' %}pill-buy{% elif s.signal == 'Avoid' %}pill-avoid{% else %}pill-hold{% endif %}">{{ s.signal }}</span></td>
<td>{{ "%.0f"|format(s.rsi) if s.rsi == s.rsi else "—" }}</td>
<td>{{ "%+.1f%%"|format(s.pct_20d * 100) if s.pct_20d == s.pct_20d else "—" }}</td>
<td>{{ "%+.1f%%"|format(s.pct_50d * 100) if s.pct_50d == s.pct_50d else "—" }}</td>
<td>{{ "%.1fx"|format(s.vol_ratio) if s.vol_ratio == s.vol_ratio else "—" }}</td>
<td><span class="pill {% if s.timing_tag == 'Good entry' %}pill-buy{% elif s.timing_tag == 'Oversold' %}pill-hold{% elif s.timing_tag == 'Extended' %}pill-avoid{% else %}pill-hold{% endif %}">{{ s.timing_tag }}</span></td>
</tr>
{% endfor %}
</tbody>
</table>
<div class="legend">RSI(14): &lt;30 oversold, &gt;70 overbought. vs 20d/50d: distance from moving average. Vol: 5-day avg volume ÷ 20-day avg (>1 = rising interest).</div>
{% endif %}

{% if dd.stocks %}
<h3 style="margin-top:1.5rem;">Top Holdings in {{ dd.top_sub }}
{% for s in dd.subsectors %}{% if s.ticker == dd.top_sub %} ({{ s.name }}){% endif %}{% endfor %}
</h3>
<table class="data" style="margin-top:.5rem;">
<thead><tr>
<th>Ticker</th><th>Name</th><th data-sort="num">Weight</th>
<th data-sort="num">Composite</th><th>Signal</th>
<th data-sort="num">RSI</th><th data-sort="num">vs 20d</th><th data-sort="num">vs 50d</th>
<th data-sort="num">Vol</th><th>Entry Timing</th>
</tr></thead>
<tbody>
{% for st in dd.stocks %}
<tr>
<td><strong>{{ st.ticker }}</strong></td>
<td>{{ st.name }}</td>
<td>{{ "%.1f%%"|format(st.weight) }}</td>
<td style="font-weight:600; color:{% if st.signal == 'Buy' %}#27ae60{% elif st.signal == 'Avoid' %}#c0392b{% else %}inherit{% endif %};">{{ "%.1f"|format(st.composite) }}</td>
<td><span class="pill {% if st.signal == 'Buy' %}pill-buy{% elif st.signal == 'Avoid' %}pill-avoid{% else %}pill-hold{% endif %}">{{ st.signal }}</span></td>
<td>{{ "%.0f"|format(st.rsi) if st.rsi == st.rsi else "—" }}</td>
<td>{{ "%+.1f%%"|format(st.pct_20d * 100) if st.pct_20d == st.pct_20d else "—" }}</td>
<td>{{ "%+.1f%%"|format(st.pct_50d * 100) if st.pct_50d == st.pct_50d else "—" }}</td>
<td>{{ "%.1fx"|format(st.vol_ratio) if st.vol_ratio == st.vol_ratio else "—" }}</td>
<td><span class="pill {% if st.timing_tag == 'Good entry' %}pill-buy{% elif st.timing_tag == 'Oversold' %}pill-hold{% elif st.timing_tag == 'Extended' %}pill-avoid{% else %}pill-hold{% endif %}">{{ st.timing_tag }}</span></td>
</tr>
{% endfor %}
</tbody>
</table>
<div class="legend">RSI(14): &lt;30 oversold, &gt;70 overbought. vs 20d/50d: distance from moving average. Vol: 5-day avg volume ÷ 20-day avg (>1 = rising interest).</div>
{% endif %}
</section>
{% endfor %}
{% endif %}

</body>
</html>
"""


def _heatmap_color(value: float):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "#f0f3f7", "#888"
    t = max(-1.0, min(1.0, value / 0.04))
    if t >= 0:
        r = int(255 - t * (255 - 30))
        g = int(255 - t * (255 - 132))
        b = int(255 - t * (255 - 73))
    else:
        r = int(255 - (-t) * (255 - 192))
        g = int(255 - (-t) * (255 - 57))
        b = int(255 - (-t) * (255 - 43))
    text = "#ffffff" if abs(t) > 0.55 else "#1A1A2A"
    return f"#{r:02x}{g:02x}{b:02x}", text


def _equity_svg(backtest_df: pd.DataFrame, width: int = 1080, height: int = 280) -> str:
    """Inline SVG line chart for strategy vs SPY equity curves."""
    if backtest_df is None or backtest_df.empty:
        return ""

    pad_l, pad_r, pad_t, pad_b = 50, 20, 20, 30
    plot_w = width - pad_l - pad_r
    plot_h = height - pad_t - pad_b

    dates = pd.to_datetime(backtest_df.index)
    x_min = dates.min().value
    x_max = dates.max().value
    y_min = float(min(backtest_df["strategy_equity"].min(), backtest_df["spy_equity"].min()))
    y_max = float(max(backtest_df["strategy_equity"].max(), backtest_df["spy_equity"].max()))
    y_min = min(y_min, 1.0) * 0.95
    y_max = max(y_max, 1.0) * 1.05

    def x(t): return pad_l + (t.value - x_min) / max(1, x_max - x_min) * plot_w
    def y(v): return pad_t + (1 - (v - y_min) / max(1e-9, y_max - y_min)) * plot_h

    def path(col):
        pts = [f"{x(d):.1f},{y(v):.1f}" for d, v in zip(dates, backtest_df[col])]
        return "M " + " L ".join(pts)

    # Y axis ticks (4)
    y_ticks = []
    for k in range(5):
        v = y_min + (y_max - y_min) * k / 4
        y_ticks.append((v, y(v)))

    # X axis: pick year markers
    yrs = sorted({d.year for d in dates})
    x_ticks = []
    for yr in yrs[::max(1, len(yrs)//6)]:
        d = pd.Timestamp(year=yr, month=1, day=1)
        if d.value < x_min or d.value > x_max:
            continue
        x_ticks.append((yr, x(d)))

    svg = [f'<svg width="100%" viewBox="0 0 {width} {height}" xmlns="http://www.w3.org/2000/svg" style="display:block;">']
    svg.append(f'<rect x="0" y="0" width="{width}" height="{height}" fill="white"/>')
    # gridlines
    for v, py in y_ticks:
        svg.append(f'<line x1="{pad_l}" y1="{py:.1f}" x2="{pad_l+plot_w}" y2="{py:.1f}" stroke="#eef2f6" stroke-width="1"/>')
        svg.append(f'<text x="{pad_l-6}" y="{py+4:.1f}" font-size="11" fill="#777" text-anchor="end">{v:.2f}</text>')
    for yr, px in x_ticks:
        svg.append(f'<text x="{px:.1f}" y="{height-8}" font-size="11" fill="#777" text-anchor="middle">{yr}</text>')

    svg.append(f'<path d="{path("spy_equity")}" fill="none" stroke="#1E3A5F" stroke-width="2" opacity="0.6"/>')
    svg.append(f'<path d="{path("strategy_equity")}" fill="none" stroke="#2E86C1" stroke-width="2.5"/>')

    # Legend
    lg_x, lg_y = pad_l + 10, pad_t + 14
    svg.append(f'<rect x="{lg_x}" y="{lg_y-10}" width="14" height="3" fill="#2E86C1"/>')
    svg.append(f'<text x="{lg_x+20}" y="{lg_y-2}" font-size="12" fill="#1A1A2A">Strategy</text>')
    svg.append(f'<rect x="{lg_x+90}" y="{lg_y-10}" width="14" height="3" fill="#1E3A5F" opacity="0.6"/>')
    svg.append(f'<text x="{lg_x+110}" y="{lg_y-2}" font-size="12" fill="#1A1A2A">SPY</text>')
    svg.append('</svg>')
    return "\n".join(svg)


def write_html(rows: List[Dict],
               cycle_info: Dict,
               heatmap_df: pd.DataFrame,
               vintage_info: Dict,
               out_path: str,
               backtest_df: Optional[pd.DataFrame] = None,
               backtest_summary: Optional[Dict] = None,
               drilldown_data: Optional[Dict] = None) -> str:
    from jinja2 import Template

    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    heatmap = []
    for tk, ser in heatmap_df.iterrows():
        cells = []
        for m in range(1, 13):
            try:
                v = float(ser.get(m))
            except Exception:
                v = float("nan")
            color, text_color = _heatmap_color(v)
            label = "—" if pd.isna(v) else f"{v*100:+.1f}%"
            cells.append({"label": label, "color": color, "text_color": text_color})
        heatmap.append((tk, cells))

    rs_rows = sorted(rows, key=lambda r: r["rs_3m"], reverse=True)
    max_abs = max((abs(r["rs_3m"]) for r in rs_rows), default=0.05) or 0.01
    for r in rs_rows:
        r["bar_pct"] = min(50.0, abs(r["rs_3m"]) / max_abs * 50.0)

    inp = cycle_info["inputs"]
    fmt = lambda v, suf="": "—" if v is None else f"{v:.2f}{suf}"
    spread = inp.get("spread")
    spread_txt = "—" if spread is None else f"{spread:+.2f}"
    s6 = inp.get("spread_6m_chg")
    spread_6m_txt = "—" if s6 is None else f"{s6:+.2f}"
    iy = inp.get("INDPRO_YoY")
    indpro_yoy_txt = "—" if iy is None else f"{iy:+.1f}"
    i3 = inp.get("INDPRO_3m_chg")
    indpro_3m_txt = "—" if i3 is None else f"{i3:+.1f}"

    s = backtest_summary or {}
    has_bt = backtest_df is not None and not backtest_df.empty
    bt_warn = bool(has_bt and not s.get("beats_spy_net"))

    tpl = Template(HTML_TEMPLATE)
    html = tpl.render(
        run_date      = date.today().strftime("%B %d, %Y"),
        run_date_long = date.today().strftime("%B %d, %Y"),
        current_year  = date.today().year,
        generated_at  = vintage_info.get("generated_at", ""),
        prices_through= vintage_info.get("prices_through", ""),
        fred_vintage  = vintage_info.get("fred_vintage", ""),

        phase           = cycle_info["phase"],
        override_active = cycle_info.get("override_active", False),
        phase_why       = cycle_info["why"],

        dgs10        = fmt(inp.get("DGS10")),
        dgs2         = fmt(inp.get("DGS2")),
        spread       = spread_txt,
        spread_6m_chg= spread_6m_txt,
        indpro_yoy   = indpro_yoy_txt,
        indpro_3m_chg= indpro_3m_txt,
        indpro_date  = str(inp.get("INDPRO_date") or "—"),
        dgs10_date   = str(inp.get("DGS10_date") or "—"),

        rows         = rows,
        rs_rows      = rs_rows,
        months       = months,
        heatmap      = heatmap,
        min_years    = config.SEASONALITY_MIN_YEARS,
        trust_yrs    = config.SEASONALITY_TRUST_YEARS,
        w_season     = f"{config.WEIGHTS.seasonality:.0%}",
        w_cycle      = f"{config.WEIGHTS.cycle_fit:.0%}",
        w_rs         = f"{config.WEIGHTS.rel_strength:.0%}",

        has_backtest    = has_bt,
        backtest_years  = config.BACKTEST_YEARS,
        backtest_warning= bt_warn,
        trade_cost_bps  = f"{config.TRADE_COST_BPS:.0f}",
        min_score       = f"{config.MIN_SCORE_TO_HOLD:.0f}",

        strategy_cum    = _pct(s.get("strategy_cum")),
        spy_cum         = _pct(s.get("spy_cum")),
        strategy_cagr   = _pct(s.get("strategy_cagr")),
        spy_cagr        = _pct(s.get("spy_cagr")),
        strategy_sharpe = _num(s.get("strategy_sharpe")),
        spy_sharpe      = _num(s.get("spy_sharpe")),
        strategy_maxdd  = _pct(s.get("strategy_maxdd")),
        spy_maxdd       = _pct(s.get("spy_maxdd")),
        equity_svg      = _equity_svg(backtest_df) if has_bt else "",

        taxable_account = config.TAXABLE_ACCOUNT,

        drilldown     = drilldown_data or {},
        sector_names  = config.SECTORS,
    )

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write(html)
    return out_path
