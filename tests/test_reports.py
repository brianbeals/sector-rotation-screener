from __future__ import annotations

from datetime import date

import pandas as pd
from openpyxl import load_workbook

import report


def sample_row():
    return {
        "ticker": "XLK",
        "name": "Technology",
        "last_price": 200.0,
        "pct_from_52w_high": -0.05,
        "seasonality_score": 70.0,
        "seasonality_n": 12,
        "seasonality_thin": False,
        "cycle_fit_score": 100.0,
        "rs_score": 65.0,
        "rs_1m": 0.01,
        "rs_3m": 0.03,
        "rs_6m": 0.05,
        "composite": 76.0,
        "signal": "Buy",
    }


def cycle_info():
    return {
        "phase": "Mid-cycle",
        "algo_phase": "Mid-cycle",
        "override_active": False,
        "why": "Synthetic test.",
        "inputs": {
            "DGS10": 4.0,
            "DGS2": 3.5,
            "spread": 0.5,
            "spread_6m_chg": 0.1,
            "INDPRO_YoY": 2.0,
            "INDPRO_3m_chg": 0.2,
            "DGS10_date": date(2024, 1, 31),
            "DGS2_date": date(2024, 1, 31),
            "INDPRO_date": date(2024, 1, 31),
        },
    }


def vintage_info():
    return {
        "generated_at": "2024-02-01 10:00",
        "prices_through": "2024-01-31",
        "fred_vintage": "2024-01-31",
    }


def test_excel_report_contains_expected_sheets(tmp_path):
    out = tmp_path / "report.xlsx"
    report.write_excel([sample_row()], cycle_info(), vintage_info(), str(out))
    workbook = load_workbook(out)
    assert workbook.sheetnames == ["Sector Screen", "Cycle Context"]
    assert workbook["Sector Screen"]["A5"].value == "XLK"
    assert workbook["Cycle Context"]["B3"].value == "Mid-cycle"


def test_html_report_contains_core_sections(tmp_path):
    out = tmp_path / "report.html"
    heatmap = pd.DataFrame([[0.01] * 12], index=["XLK"], columns=range(1, 13))
    report.write_html(
        [sample_row()],
        cycle_info(),
        heatmap,
        vintage_info(),
        str(out),
    )
    html = out.read_text()
    assert "Sector Rotation Screen" in html
    assert "Technology" in html
    assert "Mid-cycle" in html


def test_html_report_shows_underperformance_warning(tmp_path):
    out = tmp_path / "report.html"
    heatmap = pd.DataFrame([[0.01] * 12], index=["XLK"], columns=range(1, 13))
    index = pd.Index([date(2024, 1, 31), date(2024, 2, 29)], name="date")
    backtest_df = pd.DataFrame(
        {
            "strategy_equity": [1.0, 0.95],
            "spy_equity": [1.0, 1.05],
        },
        index=index,
    )
    report.write_html(
        [sample_row()],
        cycle_info(),
        heatmap,
        vintage_info(),
        str(out),
        backtest_df=backtest_df,
        backtest_summary={
            "beats_spy_net": False,
            "strategy_cum": -0.05,
            "spy_cum": 0.05,
        },
    )
    assert "Strategy did NOT beat SPY" in out.read_text()
